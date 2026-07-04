# -*- coding: utf-8 -*-
"""
====================================================================
  SEGUIMENT PRESSUPOSTARI D'INGRESSOS · App web per a ens locals
====================================================================
L'ajuntament puja el seu Excel de seguiment (format ABSIS amb la fulla
d'històric multi-any) i l'aplicació genera l'Excel d'anàlisi validat:

  · PARAMETRES              llindars configurables (inclòs Execució alta 95%)
  · ANALISI_INGRESSOS_2026  taula principal amb fórmules vives:
                            execució acumulada, % real, % esperat (perfil),
                            desviació (pp), ràtio, semàfor d'anomalia i
                            columnes de metodologia (selecció per AIC)
  · RESUM_ALERTES           recompte automàtic per nivell i per model
  · MATRIU_RISCOS           taules teòriques + detecció d'anomalies
                            sobre les dades concretes

Metodologia:
  - Acumulat NET: les sèries es netegen perquè mai decreixin (neutralitza
    devolucions/anul·lacions d'exercicis tancats i regularitzacions de padró).
  - Projecció per aplicació triant el millor model entre ETS (Holt-Winters)
    i SARIMA pel criteri d'Akaike (AIC mínim); fallback estacional simple.
  - Perfil històric esperat amb filtrat d'anys atípics (MAD + pandèmia 2020).

Executable a Google Colab (vegeu el notebook adjunt) o desplegable a
Streamlit Community Cloud per obtenir una URL pública permanent.
"""

import io
import warnings
from collections import Counter

import numpy as np
import pandas as pd
import streamlit as st

warnings.filterwarnings("ignore")

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    from statsmodels.tsa.statespace.sarimax import SARIMAX
    STATSMODELS_OK = True
except Exception:
    STATSMODELS_OK = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.properties import PageSetupProperties

st.set_page_config(page_title="Seguiment Pressupostari · Àrea de Suport Econòmic",
                   page_icon="📊", layout="centered")

# ----------------------------- ESTIL -----------------------------
ESTIL = """
<style>
.block-container { max-width: 820px; padding-top: 2.2rem; padding-bottom: 3rem; }
h1, h2, h3 { letter-spacing: -0.01em; }
.capcalera {
    border-left: 4px solid #1F3864;
    padding: 0.15rem 0 0.15rem 1rem;
    margin-bottom: 0.4rem;
}
.capcalera .area { font-size: 1.55rem; font-weight: 700; color: #1F3864; line-height: 1.15; }
.capcalera .servei { font-size: 0.95rem; color: #5A6472; margin-top: 0.15rem; }
.contacte { font-size: 0.82rem; color: #6B7280; margin-top: 0.3rem; line-height: 1.5; }
.contacte a { color: #2E5A9E; text-decoration: none; }
.contacte a:hover { text-decoration: underline; }
.pas { font-size: 0.78rem; font-weight: 600; letter-spacing: 0.06em;
       text-transform: uppercase; color: #2E5A9E; margin: 0.2rem 0 0.1rem; }
div.stButton > button[kind="primary"] {
    width: 100%; border-radius: 8px; font-weight: 600; padding: 0.55rem 0;
}
div.stDownloadButton > button {
    width: 100%; border-radius: 8px; font-weight: 600;
    background: #1F7A54; color: #fff; border: none; padding: 0.6rem 0;
}
div.stDownloadButton > button:hover { background: #17603f; color:#fff; }
.peu { font-size: 0.75rem; color: #9AA3AF; text-align: center;
       margin-top: 2.5rem; border-top: 1px solid #E5E7EB; padding-top: 0.8rem; }
</style>
"""


def capcalera():
    st.markdown(ESTIL, unsafe_allow_html=True)
    st.markdown(
        """
        <div class="capcalera">
            <div class="area">Àrea de Suport Econòmic</div>
            <div class="servei">Servei de Concertació i Assistència al Municipi</div>
        </div>
        <div class="contacte">
            Tel. 977 296 671 &nbsp;·&nbsp;
            <a href="https://www.google.com/maps/search/Pere+Martell,+2+%7C+43001+Tarragona"
               target="_blank">Pere Martell, 2 · 43001 Tarragona</a>
            &nbsp;·&nbsp; <a href="https://www.dipta.cat" target="_blank">www.dipta.cat</a>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.divider()

MESOS = ["Gen", "Feb", "Mar", "Abr", "Mai", "Jun",
         "Jul", "Ago", "Set", "Oct", "Nov", "Des"]
MESOS_CAT = {
    "gener": 0, "febrer": 1, "març": 2, "marc": 2, "abril": 3,
    "maig": 4, "juny": 5, "juliol": 6, "agost": 7, "setembre": 8,
    "octubre": 9, "novembre": 10, "desembre": 11,
}
ANY_PROJ = 2026  # any a projectar (configurable a la barra lateral)


# ==================================================================
#  1. LECTURA DE L'HISTÒRIC (format pivot ABSIS)
# ==================================================================
def llistar_fulles(xls_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xls_bytes), read_only=True, data_only=True)
    fulles = wb.sheetnames
    wb.close()
    return fulles


def llegir_historic_pivot(xls_bytes, nom_fulla, magnitud_col=6):
    """
    Llegeix la fulla d'històric en format PIVOT d'ABSIS
    ('1_1_1 HISTÒRIC INGRESSOS' o equivalent):
      - Una fila conté els blocs 'Exercici N-X: YYYY' (>= 3 blocs).
      - Cada bloc ocupa 10 columnes; dins el bloc, offsets:
          +2 econòmica, +3 descripció, +4 Mes,
          +5 prev.inicial, +6 prev.definitiva, +7 DRN, +8 recaptació
      - Per aplicació: 12 files de mes (desordenades) + fila 'Total'.

    magnitud_col: 4=prev.inicial, 5=prev.definitiva, 6=DRN, 7=recaptació.
    Retorna DataFrame llarg [aplicacio, descripcio, any, mes(0-11), valor].
    """
    wb = openpyxl.load_workbook(io.BytesIO(xls_bytes), read_only=True, data_only=True)
    ws = wb[nom_fulla]

    fila_anys, row_anys = None, None
    for r in range(1, 40):
        row = [c.value for c in ws[r]]
        matches = [v for v in row if isinstance(v, str) and "Exercici" in v
                   and ":" in v and any(ch.isdigit() for ch in v.split(":")[-1])]
        if len(matches) >= 3:
            fila_anys, row_anys = r, row
            break
    if fila_anys is None:
        wb.close()
        raise ValueError("No s'ha trobat la fila de blocs d'anys a l'històric.")

    blocs = []
    for j, v in enumerate(row_anys):
        if isinstance(v, str) and "Exercici" in v and ":" in v:
            try:
                blocs.append((j, int(str(v).split(":")[-1].strip())))
            except ValueError:
                pass

    fila_sub = None
    for r in range(fila_anys, fila_anys + 12):
        row = [c.value for c in ws[r]]
        if any(isinstance(v, str) and v.strip().lower() == "mes" for v in row if v):
            fila_sub = r
            break
    if fila_sub is None:
        fila_sub = fila_anys + 5

    OFF_ECON, OFF_DESC, OFF_MES = 2, 3, 4
    OFF_VAL = magnitud_col + 1

    totes = [list(r) for r in ws.iter_rows(min_row=fila_sub + 1,
                                           max_row=ws.max_row, values_only=True)]
    wb.close()

    dades = []
    for col0, any_ in blocs:
        econ_actual, desc_actual = None, None
        for fila in totes:
            def get(off):
                idx = col0 + off
                return fila[idx] if 0 <= idx < len(fila) else None

            econ, desc = get(OFF_ECON), get(OFF_DESC)
            mes_raw, val = get(OFF_MES), get(OFF_VAL)

            if econ is not None and not (isinstance(econ, str) and str(econ).startswith("Total")):
                ecs = str(econ).strip()
                if ecs.isdigit() and len(ecs) >= 3:
                    econ_actual = ecs
                    if desc is not None and str(desc).strip() and not str(desc).startswith("Total"):
                        desc_actual = str(desc).strip()

            if mes_raw is not None and econ_actual is not None:
                mes_key = str(mes_raw).strip().lower()
                if mes_key in MESOS_CAT:
                    try:
                        v = float(val) if val not in (None, "") else 0.0
                    except (ValueError, TypeError):
                        v = 0.0
                    dades.append({"aplicacio": econ_actual,
                                  "descripcio": desc_actual or "",
                                  "any": any_, "mes": MESOS_CAT[mes_key], "valor": v})

    df = pd.DataFrame(dades)
    if df.empty:
        raise ValueError("No s'han pogut extreure dades mensuals de l'històric.")
    return df.groupby(["aplicacio", "descripcio", "any", "mes"], as_index=False)["valor"].sum()


# ==================================================================
#  2. MOTOR DE PROJECCIÓ (selecció per AIC) + PERFIL FILTRAT
# ==================================================================
def _fallback(serie):
    serie = np.asarray(serie, float)
    n = len(serie)
    if n >= 12:
        n_anys = n // 12
        mat = serie[-n_anys * 12:].reshape(n_anys, 12)
        base = mat.mean(axis=0)
        if n_anys >= 2:
            tend = (mat[-1] - mat[0]) / max(n_anys - 1, 1)
            base = base + tend * 0.5
        return np.clip(base, 0, None)
    return np.repeat(serie[-1] if n else 0.0, 12)


def millor_projeccio_detall(serie_mensual):
    """Prova ETS i SARIMA; tria el d'AIC mínim. Retorna (fc12, model, criteri)."""
    serie = pd.Series(np.asarray(serie_mensual, float)).fillna(0.0)
    n = len(serie)
    if not STATSMODELS_OK or n < 24 or serie.sum() == 0:
        return _fallback(serie), "Estacional-simple", \
            "Sèrie curta o sense prou història: perfil estacional mitjà (sense AIC)."

    resultats, aics = {}, {}
    try:
        m = ExponentialSmoothing(serie, trend="add", seasonal="add",
                                 seasonal_periods=12,
                                 initialization_method="estimated").fit(optimized=True)
        fc = np.asarray(m.forecast(12), float)
        if np.all(np.isfinite(fc)):
            resultats["ETS (Holt-Winters)"] = fc
            aics["ETS (Holt-Winters)"] = round(float(m.aic), 1)
    except Exception:
        pass
    try:
        m = SARIMAX(serie, order=(1, 1, 1), seasonal_order=(1, 1, 0, 12),
                    enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
        fc = np.asarray(m.get_forecast(12).predicted_mean, float)
        if np.all(np.isfinite(fc)):
            resultats["SARIMA"] = fc
            aics["SARIMA"] = round(float(m.aic), 1)
    except Exception:
        pass

    if not resultats:
        return _fallback(serie), "Estacional-simple", \
            "Cap model estadístic ha convergit: perfil estacional mitjà."

    nom = min(aics, key=aics.get)
    fc = np.clip(resultats[nom], 0, None)
    if len(aics) == 2:
        criteri = (f"AIC més baix: {nom} (ETS={aics.get('ETS (Holt-Winters)')} "
                   f"vs SARIMA={aics.get('SARIMA')})")
    else:
        criteri = f"Únic model convergent: {nom}"
    return fc, nom, criteri


def perfil_filtrat(g):
    """Perfil % acumulat mitjà amb filtrat d'anys atípics (MAD + 2020)."""
    piv = g.pivot_table(index="any", columns="mes", values="valor",
                        aggfunc="sum").reindex(columns=range(12))
    piv = piv.ffill(axis=1)
    totals = piv.iloc[:, -1].astype(float)
    totals = totals[totals > 0]
    if len(totals) == 0:
        return np.zeros(12)
    med = totals.median()
    mad = (totals - med).abs().median() or 1.0
    z = 0.6745 * (totals - med).abs() / mad
    exclosos = list(totals.index[z > 3.5])
    if 2020 in totals.index and len(totals) - len(set(exclosos + [2020])) >= 3:
        if 2020 not in exclosos:
            exclosos.append(2020)
    anys_ok = [a for a in totals.index if a not in exclosos]
    if len(anys_ok) < 2:
        anys_ok = list(totals.index)
    piv_ok = piv.loc[anys_ok]
    tot_ok = piv_ok.iloc[:, -1].replace(0, np.nan)
    pct = piv_ok.div(tot_ok, axis=0).mean(axis=0)
    return pct.reindex(range(12)).ffill().fillna(0).to_numpy()


def processar(hist, any_final):
    """
    Pipeline complet: neteja (cummax), projecció AIC, perfils, crèdit ref.
    Retorna dict d'aplicacions -> {desc, proj[12], perfil[12], credit, model, aic}
    """
    hist = hist[hist["any"] <= any_final].copy()
    # ACUMULAT NET: mai decreixent dins de cada any
    hist = hist.sort_values(["aplicacio", "any", "mes"])
    hist["valor"] = hist.groupby(["aplicacio", "any"])["valor"].cummax()

    resultat = {}
    apps = [a for a, v in hist.groupby("aplicacio")["valor"].sum().items() if v > 0]
    for i, apl in enumerate(apps):
        g = hist[hist["aplicacio"] == apl].sort_values(["any", "mes"])
        fc, model, criteri = millor_projeccio_detall(g["valor"].to_numpy())
        fc = np.maximum.accumulate(np.clip(np.asarray(fc, float), 0, None))
        piv = g.pivot_table(index="any", columns="mes", values="valor", aggfunc="sum")
        tt = piv.ffill(axis=1).iloc[:, -1]
        tt = tt[tt > 0]
        resultat[apl] = {
            "desc": g["descripcio"].iloc[0] if not g["descripcio"].isna().all() else "",
            "proj": fc,
            "perfil": perfil_filtrat(g),
            "credit": float(tt.tail(3).mean()) if len(tt) else float(fc[-1]),
            "model": model,
            "aic": criteri,
        }
        yield i + 1, len(apps), apl, resultat  # progrés per a la UI


# ==================================================================
#  3. DETECCIÓ D'ANOMALIES (per a MATRIU_RISCOS)
# ==================================================================
def detectar(resultat):
    det = {1: [], 2: [], 3: [], 4: []}
    for apl, d in resultat.items():
        acum, perfil = d["proj"], d["perfil"]
        total = acum[-1]
        if total <= 0:
            continue
        inc = np.clip(np.diff(acum, prepend=0.0), 0, None)
        inc_pos = inc[inc > 0]
        mitjana_inc = inc_pos.mean() if len(inc_pos) else 0
        inc_perfil = np.clip(np.diff(perfil, prepend=0.0), 0, None) * total

        for m in range(12):
            esperat = inc_perfil[m]
            if inc[m] > 2 * max(esperat, total * 0.02) and inc[m] > total * 0.15:
                det[1].append((apl, d["desc"], MESOS[m],
                               f"{inc[m]:,.0f} € (esperat {esperat:,.0f} €)"))
                break
        pct9 = 100 * acum[8] / total
        if pct9 < 30:
            det[2].append((apl, d["desc"], "Set", f"{pct9:.1f}% acumulat (llindar 30%)"))
        if perfil[8] > 0.05:
            ratio = (acum[8] / total) / perfil[8]
            if ratio < 0.75:
                det[3].append((apl, d["desc"], "Set",
                               f"{ratio*100:.0f}% del perfil (esperat {perfil[8]*100:.0f}%)"))
        for i in range(1, len(inc) - 1):
            a, b = inc[i], inc[i + 1]
            if a > 0 and b > 0 and abs(a - b) <= 0.01 * max(a, b) \
                    and a >= mitjana_inc and mitjana_inc > 0:
                det[4].append((apl, d["desc"], f"{MESOS[i]}–{MESOS[i+1]}",
                               f"{a:,.0f} € repetit"))
                break
    return det


# ==================================================================
#  4. CONSTRUCCIÓ DE L'EXCEL DE SORTIDA (format validat, 4 pestanyes)
# ==================================================================
def construir_excel(resultat, det, any_proj=ANY_PROJ, mes_defecte=5):
    FONT = "Calibri"
    blau_fosc = PatternFill("solid", fgColor="1F3864")
    blau_seco = PatternFill("solid", fgColor="2E4C7E")
    blau_clar = PatternFill("solid", fgColor="EAF0F8")
    groc = PatternFill("solid", fgColor="FFC000")
    vermell = PatternFill("solid", fgColor="C00000")
    verd = PatternFill("solid", fgColor="1F7A54")
    taronja = PatternFill("solid", fgColor="ED7D31")
    blau_info = PatternFill("solid", fgColor="5B9BD5")
    blanc_b = Font(bold=True, color="FFFFFF", name=FONT, size=11)
    negre = Font(name=FONT, size=10)
    negre_b = Font(bold=True, name=FONT, size=10)
    blau_txt = Font(name=FONT, size=10, color="0000CC")
    gris_txt = Font(name=FONT, size=8, color="999999")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    vora = Border(thin, thin, thin, thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # -------------------- PARAMETRES --------------------
    wsp = wb.create_sheet("PARAMETRES")
    wsp["B1"] = "PARÀMETRES DEL MODEL"
    wsp["B1"].font = Font(bold=True, size=13, color="1F3864", name=FONT)
    par = [
        ("B7", "Llindars d'Alerta (configurables)", negre_b, None),
        ("B8", "Nivell", blanc_b, blau_fosc), ("C8", "Llindar (ràtio)", blanc_b, blau_fosc),
        ("D8", "Descripció", blanc_b, blau_fosc), ("E8", "Acció recomanada", blanc_b, blau_fosc),
        ("B9", "CRÍTIC (sobre)", negre, None), ("C9", 0.4, blau_txt, None),
        ("D9", "Execució >40% per sobre l'esperada", negre, None),
        ("E9", "Notificació immediata Interventor + Regidor", negre, None),
        ("B10", "AVÍS (sobre)", negre, None), ("C10", 0.2, blau_txt, None),
        ("D10", "Execució entre +20% i +40% sobre l'esperada", negre, None),
        ("E10", "Revisar en propera reunió de seguiment", negre, None),
        ("B11", "AVÍS (sota)", negre, None), ("C11", -0.4, blau_txt, None),
        ("D11", "Execució entre -40% i -20% sota l'esperada", negre, None),
        ("E11", "Revisar en propera reunió de seguiment", negre, None),
        ("B12", "CRÍTIC (sota)", negre, None), ("C12", -0.6, blau_txt, None),
        ("D12", "Execució >60% per sota l'esperada", negre, None),
        ("E12", "Notificació immediata Interventor + Regidor", negre, None),
        ("B13", "PENDENT", negre, None), ("C13", 0.05, blau_txt, None),
        ("D13", "Execució 0% quan perfil espera >5%", negre, None),
        ("E13", "Comprovar bloqueig administratiu", negre, None),
        ("B14", "EXECUCIÓ ALTA", negre, None), ("C14", 0.95, blau_txt, None),
        ("D14", "Si % Real ≥ aquest valor, no s'apliquen alertes per sota", negre, None),
        ("E14", "L'execució ja és pràcticament completa", negre, None),
        ("B15", "Configuració Temporal", negre_b, None),
        ("B16", "Any analitzat", negre, None), ("C16", any_proj, blau_txt, None),
        ("B17", "Mes actual", negre, None), ("C17", mes_defecte, blau_txt, None),
        ("D17", "Mes de referència (1=Gen … 12=Des)", negre, None),
    ]
    for ref, val, fnt, fill in par:
        c = wsp[ref]; c.value = val; c.font = fnt
        if fill:
            c.fill = fill
        c.alignment = left if isinstance(val, str) else center
    for col, w in zip("BCDE", [18, 14, 46, 42]):
        wsp.column_dimensions[col].width = w

    # -------------------- ANALISI --------------------
    nom_full = f"ANALISI_INGRESSOS_{any_proj}"
    ws = wb.create_sheet(nom_full)
    ws["A1"] = f"SEGUIMENT PRESSUPOSTARI – INGRESSOS (Projecció {any_proj})"
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=13, color="1F3864", name=FONT)
    ws["A2"] = "Any:"; ws["A2"].font = negre_b
    ws["B2"] = "=PARAMETRES!C16"; ws["B2"].font = blau_txt; ws["B2"].alignment = center
    ws["C2"] = "Mes analitzat:"; ws["C2"].font = negre_b; ws["C2"].alignment = right
    ws["D2"] = "=PARAMETRES!C17"; ws["D2"].font = blau_txt; ws["D2"].alignment = center

    seccions = [("A3", "IDENTIFICACIÓ", "A3:D3"),
                ("E3", "EXECUCIÓ MENSUAL (acumulada en €)", "E3:P3"),
                ("Q3", "EXECUCIÓ ACUMULADA", "Q3:R3"),
                ("S3", "ANÀLISI DESVIACIÓ", "S3:U3"),
                ("V3", "ALERTA", "V3:W3"),
                ("AN3", "MODEL DE PROJECCIÓ (AIC)", "AN3:AO3")]
    for ref, txt, rang in seccions:
        ws[ref] = txt; ws.merge_cells(rang)
        ws[ref].fill = blau_fosc; ws[ref].font = blanc_b; ws[ref].alignment = center

    caps = ["Aplicació pressupost.", "Descripció", "Crèdit inici (€)", "Crèdit def. (€)"] \
        + MESOS + ["Oblig./Drets acum. (€)", "% Real", "% Esperat (model)",
                   "Desviació (pp)", "Ràtio desv.", "Anomalia", "Justificació/Estat"]
    for i, txt in enumerate(caps, start=1):
        c = ws.cell(row=4, column=i, value=txt)
        c.fill = blau_seco; c.font = blanc_b; c.alignment = center; c.border = vora
    ws.cell(row=4, column=40, value="Metodologia").fill = blau_seco
    ws.cell(row=4, column=41, value="Selecció (AIC)").fill = blau_seco
    for c in (40, 41):
        ws.cell(4, c).font = blanc_b; ws.cell(4, c).alignment = center; ws.cell(4, c).border = vora

    r = 5
    for apl, d in resultat.items():
        p, perf = d["proj"], d["perfil"]
        ws.cell(row=r, column=1, value=apl).font = negre_b
        ws.cell(row=r, column=2, value=d["desc"]).font = negre
        ws.cell(row=r, column=3, value=round(d["credit"], 0)).font = blau_txt
        ws.cell(row=r, column=4, value=round(d["credit"], 0)).font = blau_txt
        for m in range(12):
            cc = ws.cell(row=r, column=5 + m, value=round(float(p[m]), 2))
            cc.font = blau_txt; cc.number_format = "#,##0"
            pc = ws.cell(row=r, column=25 + m, value=round(float(perf[m]), 4))
            pc.number_format = "0.0%"; pc.font = gris_txt

        ws.cell(row=r, column=17,
                value=f"=IFERROR(CHOOSE(PARAMETRES!$C$17,E{r},F{r},G{r},H{r},I{r},"
                      f"J{r},K{r},L{r},M{r},N{r},O{r},P{r}),0)")
        ws.cell(row=r, column=18, value=f"=IFERROR(Q{r}/D{r},0)")
        ws.cell(row=r, column=19,
                value=f"=IFERROR(CHOOSE(PARAMETRES!$C$17,Y{r},Z{r},AA{r},AB{r},AC{r},"
                      f"AD{r},AE{r},AF{r},AG{r},AH{r},AI{r},AJ{r}),0)")
        ws.cell(row=r, column=20, value=f"=R{r}-S{r}")
        ws.cell(row=r, column=21, value=f"=IFERROR((Q{r}-(D{r}*S{r}))/(D{r}*S{r}),0)")
        ws.cell(row=r, column=22,
                value=(f'=IFERROR(IF(T{r}>=PARAMETRES!$C$9,"🔴 CRÍTIC (sobre)",'
                       f'IF(T{r}>=PARAMETRES!$C$10,"🟡 AVÍS (sobre)",'
                       f'IF(AND(T{r}<=PARAMETRES!$C$12,R{r}<PARAMETRES!$C$14),"🔴 CRÍTIC (sota)",'
                       f'IF(AND(T{r}<=PARAMETRES!$C$11,R{r}<PARAMETRES!$C$14),"🟡 AVÍS (sota)",'
                       f'IF(AND(R{r}=0,S{r}>=PARAMETRES!$C$13),"🔵 PENDENT",'
                       f'"🟢 NORMAL"))))),"")'))
        ws.cell(row=r, column=23, value="")
        ws.cell(row=r, column=40, value=d["model"]).font = negre
        ws.cell(row=r, column=41, value=d["aic"]).font = Font(name=FONT, size=8)

        for c in range(1, 24):
            cell = ws.cell(row=r, column=c)
            cell.border = vora
            cell.alignment = left if c in (1, 2) else (center if c >= 18 else right)
            if c in (3, 4, 17):
                cell.number_format = "#,##0"
            if c in (18, 19, 20):
                cell.number_format = "0.0%"
            if c == 21:
                cell.number_format = '0.00"x"'
            if r % 2 == 0:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", None):
                    cell.fill = blau_clar
        for c in (40, 41):
            ws.cell(r, c).border = vora
            ws.cell(r, c).alignment = left
        r += 1

    n_files = r - 1
    ws.cell(row=r, column=1, value="TOTALS").font = blanc_b
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value=f"=SUM(C5:C{n_files})")
    ws.cell(row=r, column=4, value=f"=SUM(D5:D{n_files})")
    ws.cell(row=r, column=17, value=f"=SUM(Q5:Q{n_files})")
    for c in range(1, 24):
        cc = ws.cell(row=r, column=c)
        cc.border = vora; cc.fill = blau_fosc; cc.font = blanc_b
        if c in (3, 4, 17):
            cc.number_format = "#,##0"; cc.alignment = right

    amp = {"A": 20, "B": 34, "C": 15, "D": 15, "Q": 20, "R": 9, "S": 15,
           "T": 13, "U": 11, "V": 18, "W": 22, "AN": 20, "AO": 40}
    for col in "EFGHIJKLMNOP":
        amp[col] = 9
    for col, w in amp.items():
        ws.column_dimensions[col].width = w
    for c in range(25, 37):
        ws.column_dimensions[get_column_letter(c)].hidden = True
    ws.freeze_panes = "C5"

    rv = f"V5:V{n_files}"
    ws.conditional_formatting.add(rv, FormulaRule(
        formula=['ISNUMBER(SEARCH("CRÍTIC",V5))'],
        fill=PatternFill("solid", fgColor="F4B7B7"),
        font=Font(bold=True, color="9C0006", name=FONT)))
    ws.conditional_formatting.add(rv, FormulaRule(
        formula=['ISNUMBER(SEARCH("AVÍS",V5))'],
        fill=PatternFill("solid", fgColor="FFE699"),
        font=Font(bold=True, color="9C6500", name=FONT)))
    ws.conditional_formatting.add(rv, FormulaRule(
        formula=['ISNUMBER(SEARCH("PENDENT",V5))'],
        fill=PatternFill("solid", fgColor="BDD7EE"),
        font=Font(bold=True, color="1F4E78", name=FONT)))
    ws.conditional_formatting.add(rv, FormulaRule(
        formula=['ISNUMBER(SEARCH("NORMAL",V5))'],
        fill=PatternFill("solid", fgColor="C6EFCE"),
        font=Font(color="006100", name=FONT)))
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:4"

    # -------------------- RESUM_ALERTES --------------------
    wr = wb.create_sheet("RESUM_ALERTES")
    wr["B2"] = "RESUM D'ALERTES"
    wr["B2"].font = Font(bold=True, size=14, color="1F3864", name=FONT)
    wr["B4"] = "INDICADORS GLOBALS"; wr["B4"].font = negre_b
    rng = f"{nom_full}!$V$5:$V${n_files}"
    items = [
        ("B5", "🔴 Crítiques", "C5", f'=COUNTIF({rng},"*CRÍTIC*")', "FCE4E4"),
        ("B6", "🟡 Avisos", "C6", f'=COUNTIF({rng},"*AVÍS*")', "FFF4D6"),
        ("B7", "🔵 Pendents", "C7", f'=COUNTIF({rng},"*PENDENT*")', "E1EEF9"),
        ("B8", "🟢 Normals", "C8", f'=COUNTIF({rng},"*NORMAL*")', "E4F4E4"),
        ("B9", "Σ Total aplicacions", "C9",
         f"=COUNTA({nom_full}!$A$5:$A${n_files})", "F2F2F2"),
    ]
    for b, bt, cc, cf, col in items:
        wr[b] = bt; wr[b].font = negre_b
        wr[b].fill = PatternFill("solid", fgColor=col)
        wr[cc] = cf; wr[cc].font = negre_b; wr[cc].alignment = center
        wr[cc].fill = PatternFill("solid", fgColor=col)
    wr["B11"] = "MODELS DE PROJECCIÓ EMPRATS"; wr["B11"].font = negre_b
    rr = 12
    for k, v in Counter(d["model"] for d in resultat.values()).most_common():
        wr.cell(row=rr, column=2, value=k).font = negre
        wr.cell(row=rr, column=3, value=v).font = negre_b
        wr.cell(row=rr, column=3).alignment = center
        rr += 1
    wr.column_dimensions["B"].width = 26
    wr.column_dimensions["C"].width = 12

    # -------------------- MATRIU_RISCOS --------------------
    wm = wb.create_sheet("MATRIU_RISCOS")
    wm["B1"] = "MATRIU DE RISCOS I CIRCUIT D'ALERTES"
    wm["B1"].font = Font(bold=True, size=14, color="1F3864", name=FONT)

    wm["B3"] = "NIVELLS D'ALERTA"; wm.merge_cells("B3:E3")
    wm["B3"].fill = PatternFill("solid", fgColor="2E75B6")
    wm["B3"].font = blanc_b; wm["B3"].alignment = center
    for i, t in enumerate(["Nivell", "Valor", "Descripció", "Acció recomanada"]):
        c = wm.cell(row=4, column=2 + i, value=t)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.font = blanc_b; c.alignment = center; c.border = vora
    nivells = [
        ("CRÍTIC (per sobre)", "40,0%", "Execució >40% per sobre la prevista",
         "Notificació immediata Interventor + Regidor", vermell, "FFFFFF"),
        ("AVÍS (per sobre)", "20,0%", "Execució entre +20% i +40% per sobre la prevista",
         "Revisar en la següent reunió de seguiment", groc, "000000"),
        ("AVÍS (per sota)", "-40,0%", "Execució entre -40% i -20% per sota la prevista",
         "Revisar en la següent reunió de seguiment", groc, "000000"),
        ("CRÍTIC (per sota)", "-60,0%", "Execució >60% per sota la prevista",
         "Notificació immediata Interventor + Regidor", vermell, "FFFFFF"),
        ("PENDENT", "5,0%", "Execució 0% quan la previsió espera >5%",
         "Comprovar si hi ha bloqueig administratiu", groc, "000000"),
        ("NORMAL", "—", "Resta de casos", "Sense accions addicionals", verd, "FFFFFF"),
    ]
    r = 5
    for nom, val, desc, acc, fill, tc in nivells:
        wm.cell(row=r, column=2, value=nom)
        wm.cell(row=r, column=3, value=val)
        wm.cell(row=r, column=4, value=desc)
        wm.cell(row=r, column=5, value=acc)
        for c in range(2, 6):
            cell = wm.cell(row=r, column=c); cell.fill = fill
            cell.font = Font(bold=(c == 2), color=tc, name=FONT, size=10)
            cell.border = vora; cell.alignment = center if c == 3 else left
        r += 1

    r += 1
    wm.cell(row=r, column=2, value="TIPOLOGIA D'ANOMALIES (naturalesa causal)")
    wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    wm.cell(row=r, column=2).fill = blau_fosc
    wm.cell(row=r, column=2).font = blanc_b
    wm.cell(row=r, column=2).alignment = center
    r += 1
    for i, t in enumerate(["Tipus d'anomalia", "Criteri tècnic", "Implicació operativa"]):
        c = wm.cell(row=r, column=2 + i, value=t)
        c.fill = blau_fosc; c.font = blanc_b; c.alignment = center; c.border = vora
    r += 1
    anomalies = [
        ("Despesa sobrevinguda", "> 2 vegades la mitjana mensual",
         "Auditoria d'urgència de l'expedient"),
        ("Infraexecució crònica", "< 30% acumulat (mes 9)",
         "Reavaluació de la viabilitat del projecte"),
        ("Ingressos endarrerits", "< 75% del perfil esperat",
         "Accions de gestió tributària o recaptació"),
        ("Pagament duplicat", "Patró de recurrència < 30 dies",
         "Bloqueig automàtic de pagaments"),
    ]
    for i, (tip, crit, impl) in enumerate(anomalies):
        wm.cell(row=r, column=2, value=tip)
        wm.cell(row=r, column=3, value=crit)
        wm.cell(row=r, column=4, value=impl)
        fill = PatternFill("solid", fgColor="F2F2F2" if i % 2 else "E8EDF5")
        for c in range(2, 5):
            cell = wm.cell(row=r, column=c); cell.fill = fill
            cell.font = negre_b if c == 2 else negre
            cell.border = vora; cell.alignment = left if c != 3 else center
        r += 1

    r += 2
    wm.cell(row=r, column=2,
            value=f"DETECCIÓ D'ANOMALIES SOBRE LES DADES (Projecció {any_proj}, acumulat net)")
    wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wm.cell(row=r, column=2).fill = blau_fosc
    wm.cell(row=r, column=2).font = Font(bold=True, size=12, color="FFFFFF", name=FONT)
    wm.cell(row=r, column=2).alignment = center
    r += 1
    wm.cell(row=r, column=2,
            value="Nota: sèries netes de devolucions/anul·lacions d'exercicis tancats "
                  "(acumulat mai decreixent) per no distorsionar l'execució.")
    wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wm.cell(row=r, column=2).font = Font(italic=True, size=9, color="666666", name=FONT)
    r += 2

    blocs = [
        (1, "1 · Despesa sobrevinguda", "Pic d'increment > 2× l'esperat pel perfil",
         "ℹ️ INFORMATIVA en ingressos: un pic de meritació sol ser normal. Revisar només si no s'esperava.",
         blau_info, "FFFFFF"),
        (2, "2 · Infraexecució crònica", "< 30% acumulat al mes 9 (setembre)",
         "🔴 Reavaluar la viabilitat del projecte / anul·lar crèdit.", vermell, "FFFFFF"),
        (3, "3 · Ingressos endarrerits", "< 75% del perfil històric esperat",
         "🟡 Accions de gestió tributària o de recaptació.", groc, "000000"),
        (4, "4 · Pagament/ingrés duplicat", "Dos increments mensuals consecutius idèntics",
         "🟠 Verificar duplicitat i bloquejar si escau.", taronja, "FFFFFF"),
    ]
    for num, titol, criteri, implicacio, fill, tc in blocs:
        wm.cell(row=r, column=2, value=titol)
        wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = wm.cell(row=r, column=2)
        c.fill = fill; c.font = Font(bold=True, color=tc, size=11, name=FONT)
        c.alignment = left
        r += 1
        wm.cell(row=r, column=2, value=f"Criteri: {criteri}")
        wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        wm.cell(row=r, column=2).font = Font(italic=True, size=9, color="444444", name=FONT)
        r += 1
        wm.cell(row=r, column=2, value=implicacio)
        wm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        wm.cell(row=r, column=2).font = Font(size=9, color="666666", name=FONT)
        r += 1
        llista = det[num]
        if not llista:
            wm.cell(row=r, column=2, value="   — Cap aplicació detectada amb aquest criteri —")
            wm.cell(row=r, column=2).font = Font(italic=True, size=9, color="999999", name=FONT)
            r += 2
            continue
        for i, t in enumerate(["Aplicació", "Descripció", "Mes", "Valor detectat"]):
            c = wm.cell(row=r, column=2 + i, value=t)
            c.fill = blau_fosc; c.font = blanc_b; c.alignment = center; c.border = vora
        r += 1
        for apl, desc, mes, val in llista:
            wm.cell(row=r, column=2, value=apl).font = negre_b
            wm.cell(row=r, column=3, value=desc).font = negre
            wm.cell(row=r, column=4, value=mes).font = negre
            wm.cell(row=r, column=5, value=val).font = negre
            for c in range(2, 6):
                cell = wm.cell(row=r, column=c)
                cell.border = vora
                cell.alignment = left if c in (2, 3, 5) else center
            r += 1
        wm.cell(row=r, column=2, value=f"   Total detectades: {len(llista)} aplicacions")
        wm.cell(row=r, column=2).font = Font(bold=True, size=9, color="444444", name=FONT)
        r += 2
    for col, w in zip("BCDEF", [22, 40, 12, 42, 14]):
        wm.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ==================================================================
#  5. INTERFÍCIE STREAMLIT
# ==================================================================
def main():
    capcalera()

    st.title("Seguiment pressupostari d'ingressos")
    st.write("Aquesta eina projecta l'execució d'ingressos del proper exercici, "
             "detecta desviacions i genera un informe en Excel amb la matriu de "
             "riscos. Tot el càlcul es fa en aquesta pàgina i **el fitxer original "
             "no es modifica**.")

    if not STATSMODELS_OK:
        st.warning("`statsmodels` no està instal·lat: s'usarà només la projecció "
                   "estacional simple.")

    # ---- barra lateral: configuració ----
    with st.sidebar:
        st.markdown("### Configuració")
        any_proj = st.number_input("Any a projectar", 2024, 2040, ANY_PROJ)
        mes_defecte = st.slider("Mes analitzat inicial", 1, 12, 5,
                                help="Es podrà canviar després a la pestanya "
                                     "PARAMETRES de l'Excel.")
        st.divider()
        st.caption("**Llindars d'alerta** (editables a l'Excel resultant): "
                   "crític ±40/−60 pp · avís ±20/−40 pp · pendent 5% · "
                   "execució alta 95%. En canviar-los, tota la taula es recalcula.")

    # ---- pas 1: pujar fitxer ----
    st.markdown('<div class="pas">Pas 1 · Fitxer</div>', unsafe_allow_html=True)
    fitxer = st.file_uploader("Puja l'Excel de seguiment (.xlsx, format ABSIS)",
                              type=["xlsx"], label_visibility="visible")
    if fitxer is None:
        st.info("Comença pujant el teu Excel de seguiment.")
        _peu()
        st.stop()

    xls_bytes = fitxer.read()
    with st.spinner("Llegint el llibre…"):
        fulles = llistar_fulles(xls_bytes)
    st.success(f"Llibre carregat correctament · {len(fulles)} fulles.")

    # ---- pas 2: opcions ----
    st.markdown('<div class="pas">Pas 2 · Opcions</div>', unsafe_allow_html=True)
    default_idx = 0
    for i, f in enumerate(fulles):
        if "HISTÒRIC" in f.upper() and "INGRES" in f.upper():
            default_idx = i
            break
    col_a, col_b = st.columns(2)
    with col_a:
        fulla = st.selectbox("Fulla d'històric d'ingressos", fulles, index=default_idx,
                             help="Pestanya del teu Excel amb l'històric multi-any.")
    with col_b:
        magnitud = st.selectbox("Magnitud a projectar",
                                ["DRN – Drets reconeguts nets", "Recaptació neta",
                                 "Previsió definitiva", "Previsió inicial"],
                                help="DRN és l'execució d'ingressos (recomanat). "
                                     "Recaptació és el cobrat de veritat.")
    mag_map = {"Previsió inicial": 4, "Previsió definitiva": 5,
               "DRN – Drets reconeguts nets": 6, "Recaptació neta": 7}

    # ---- pas 3: generar ----
    st.markdown('<div class="pas">Pas 3 · Generar</div>', unsafe_allow_html=True)
    if not st.button("Generar l'Excel d'anàlisi", type="primary"):
        _peu()
        st.stop()

    try:
        with st.spinner("Extraient l'històric…"):
            hist = llegir_historic_pivot(xls_bytes, fulla, mag_map[magnitud])
    except Exception as e:
        st.error(f"No s'ha pogut llegir l'històric: {e}")
        st.stop()

    st.caption(f"Històric extret · {hist['aplicacio'].nunique()} aplicacions "
               f"· {hist['any'].nunique()} exercicis.")

    barra = st.progress(0.0, "Projectant…")
    resultat = {}
    for fet, total, apl, resultat in processar(hist, any_proj - 1):
        barra.progress(fet / total, f"Projectant aplicacions… {fet}/{total}")
    barra.empty()

    with st.spinner("Detectant anomalies i construint l'Excel…"):
        det = detectar(resultat)
        xlsx = construir_excel(resultat, det, any_proj, mes_defecte)

    models = Counter(d["model"] for d in resultat.values())

    # ---- resultat ----
    st.divider()
    st.markdown("#### Resultat")
    st.write(f"S'han projectat **{len(resultat)} aplicacions** amb el millor "
             "model per a cadascuna (selecció per criteri AIC):")
    m1, m2, m3 = st.columns(3)
    m1.metric("SARIMA", models.get("SARIMA", 0))
    m2.metric("ETS (Holt-Winters)", models.get("ETS (Holt-Winters)", 0))
    m3.metric("Estacional simple", models.get("Estacional-simple", 0))

    st.write("Anomalies detectades sobre la projecció:")
    a1, a2, a3, a4 = st.columns(4)
    a1.metric("Pics meritació", len(det[1]), help="Informatiu en ingressos")
    a2.metric("Infraexecució", len(det[2]))
    a3.metric("Ingr. endarrerits", len(det[3]))
    a4.metric("Possibles duplicats", len(det[4]))

    st.write("")
    st.download_button(
        "⬇  Descarregar l'Excel d'anàlisi",
        data=xlsx,
        file_name=f"ANALISI_INGRESSOS_{any_proj}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("L'Excel conté fórmules vives: canviant el mes o els llindars a la "
               "pestanya PARAMETRES, tota l'anàlisi es recalcula automàticament.")
    _peu()


def _peu():
    st.markdown(
        '<div class="peu">Àrea de Suport Econòmic · Servei de Concertació i '
        'Assistència al Municipi · Diputació de Tarragona</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
